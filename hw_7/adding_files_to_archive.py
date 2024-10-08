import os
import zipfile
import csv
import pytest
from pypdf import PdfReader
from openpyxl import load_workbook
from zipfile import ZipFile

current_dir = os.path.dirname(__file__)  # путь текущей папки
dir_path = os.path.join(current_dir, 'tmp')  # путь к папке тмп
archive_path = os.path.join(dir_path, "new_archive.zip")  # имя создаваемого архива
files_dir = [i for i in os.listdir(dir_path)]  # список файлов в папке тмп


@pytest.fixture(scope='session', autouse=True)
def create_delete_archive():
    with zipfile.ZipFile(archive_path, 'w') as zf:
        for file in files_dir:
            add_file = os.path.join(dir_path, file)
            zf.write(add_file, os.path.basename(add_file))
    yield
    os.remove(os.path.join(archive_path))


def test_create_archive():
    with ZipFile(archive_path) as zf:
        assert zf.namelist() == files_dir


def test_content_check_xlsx():
    with zipfile.ZipFile(archive_path) as zf:
        with zf.open('File1.xlsx') as xlsx_f:
            workbook = load_workbook(xlsx_f)
            sheet = workbook.active
            test_data = sheet.cell(row=3, column=2).value
            assert "Мурзик" in test_data


def test_content_check_pdf():
    with zipfile.ZipFile(archive_path) as zf:
        with zf.open('File2.pdf') as pdf_f:
            reader = PdfReader(pdf_f)
            text = reader.get_page(0).extract_text()
            assert "Автоматизация" in text


def test_content_check_csv():
    with zipfile.ZipFile(archive_path) as zf:
        with zf.open("file3.csv") as csv_f:
            content = csv_f.read().decode('utf-8')
            csvreader = list(csv.reader(content.splitlines(), delimiter=';'))
            second_row = csvreader[1]
            assert second_row[0] == '0000000000'
